import streamlit as st
import openpyxl
from copy import copy
import io
import zipfile

st.set_page_config(page_title="样式级 Excel 拆分器", layout="centered")

st.title("💎 跨境大表样式级拆分器")
st.caption("100% 保留合并单元格、字体、颜色及隐藏表结构")

uploaded_file = st.file_uploader("上传 Excel 模板", type=["xlsx"])

if uploaded_file:
    # 加载工作簿（为了获取 Sheet 名）
    wb_info = openpyxl.load_workbook(uploaded_file, read_only=True)
    sheet_names = wb_info.sheetnames

    col1, col2 = st.columns(2)
    with col1:
        target_sheet_name = st.selectbox("选择要拆分的数据页", sheet_names)
    with col2:
        chunk_size = st.number_input("每包保留的数据行数", value=2500, step=500)

    header_rows = st.number_input("表头行数 (这些行永远保留)", value=6)

    if st.button("🔥 开始完美拆分", type="primary", use_container_width=True):
        try:
            # 重新加载完整工作簿（非只读，用于操作）
            uploaded_file.seek(0)
            base_wb = openpyxl.load_workbook(uploaded_file)
            ws = base_wb[target_sheet_name]

            max_row = ws.max_row
            data_rows_count = max_row - header_rows

            if data_rows_count <= 0:
                st.error("表中没有发现可拆分的数据行，请检查表头行数设置。")
                st.stop()

            num_chunks = (data_rows_count // chunk_size) + (1 if data_rows_count % chunk_size != 0 else 0)

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                for i in range(num_chunks):
                    # 每次循环都基于原始 WB 制作一个深拷贝
                    # 这样所有 Sheet、样式、合并单元格都在
                    import copy as python_copy

                    new_wb = openpyxl.load_workbook(uploaded_file)
                    new_ws = new_wb[target_sheet_name]

                    # 计算当前块的数据范围
                    current_start = header_rows + (i * chunk_size) + 1
                    current_end = header_rows + ((i + 1) * chunk_size)

                    # 关键逻辑：删除当前块之后的所有行
                    if max_row > current_end:
                        new_ws.delete_rows(current_end + 1, max_row - current_end)

                    # 关键逻辑：删除表头之后、当前块之前的所有行
                    if current_start > header_rows + 1:
                        new_ws.delete_rows(header_rows + 1, current_start - (header_rows + 1))

                    # 将处理好的工作簿存入内存
                    output = io.BytesIO()
                    new_wb.save(output)

                    zip_file.writestr(f"Part_{i + 1}_Perfect.xlsx", output.getvalue())
                    st.write(f"✅ 已完成第 {i + 1} 个文件的样式保留处理")

            st.success(f"✨ 成功！所有文件已完美保留原表格式。")
            st.download_button(
                "📥 下载完美格式压缩包 (ZIP)",
                zip_buffer.getvalue(),
                f"Perfect_Split_{uploaded_file.name}.zip",
                use_container_width=True
            )

        except Exception as e:
            st.error(f"处理出错: {e}")